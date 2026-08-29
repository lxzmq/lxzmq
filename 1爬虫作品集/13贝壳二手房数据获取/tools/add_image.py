# load_workbook用于重新打开已经生成的Excel文件
from openpyxl import load_workbook
# Image用于把下载的房源图片插入Excel
from openpyxl.drawing.image import Image
# BytesIO用于把网络图片暂存在内存中，不需要先保存到本地
from io import BytesIO
import requests

"""
【图片插入Excel函数】
下载房源封面图片，嵌入到Excel对应行中。
参数：
file_name：目标Excel文件名。
image_urls：图片链接列表，与Excel数据行一一对应。
column：图片要插入的列，例如"G"列。
"""
def add_images_to_excel(file_name, image_urls, column):
    # 请求头模拟普通浏览器访问，降低图片服务器拒绝请求的概率
    headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
    # 打开指定的Excel文件，并把工作簿对象保存到wb变量
    wb = load_workbook(file_name)
    # active表示获取当前工作簿中默认激活的工作表
    ws = wb.active
    # 设置图片所在列的宽度为18，让图片显示区域更宽
    ws.column_dimensions[column].width = 18
    # enumerate同时取得行号和图片网址；从2开始是因为第1行是表头
    for row_num, img_url in enumerate(image_urls, 2):
        # 设置当前数据行的高度为90，使单元格能够容纳图片
        ws.row_dimensions[row_num].height = 90
        # f字符串会把列名和行号拼接起来，例如G和2会得到G2
        ws[f"{column}{row_num}"] = ""
        # 如果当前图片网址为空，就跳过本次循环
        if not img_url:
            continue
        # try中的代码可能因网络或图片格式问题报错，因此使用异常处理
        try:
            # 下载图片，将二进制内容交给BytesIO包装成内存文件
            img_data = BytesIO(requests.get(img_url, headers=headers, timeout=10).content)
            # 把内存中的图片数据转换为openpyxl可使用的图片对象
            img = Image(img_data)
            # 将图片显示尺寸统一设置为宽120像素、高90像素
            img.width, img.height = 120, 90
            # 把图片插入对应单元格位置，例如G2、G3
            ws.add_image(img, f"{column}{row_num}")
        # 捕获图片下载或插入时出现的异常，并将错误保存到变量e
        except Exception as e:
            # 某张图片失败时只打印提示，不会让整个程序停止
            print("图片插入失败:", e)
    # 保存工作簿，把插入的图片和格式真正写入Excel文件
    wb.save(file_name)